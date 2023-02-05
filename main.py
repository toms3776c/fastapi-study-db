from pathlib import Path
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse

import openpyxl
import pprint

app = FastAPI()

origins = [
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def Hello():
    return {"Hello":"World!"}


@app.post("/export-excel1")
def export_excel1():
    filename = "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1,1).value = "テスト"
    wb.save(filename)
    wb.close()
    wb = open(filename, "rb")
    
    #MIMEタイプを設定
    XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    headers = {"Content-Disposition": "attachment; filename=" + filename}
    return StreamingResponse(content=wb, media_type=XLSX_MIMETYPE, headers=headers)


@app.get("/export-excel2")
def export_excel2():
    filename = "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1,1).value = "テスト"
    wb.save(filename)
    wb.close()
    wb = open(filename, "rb")
    
    current = Path()
    file_path = current / filename
    return FileResponse(path=file_path, filename=filename)

